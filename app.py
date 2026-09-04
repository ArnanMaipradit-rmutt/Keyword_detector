import os
import io
import re
import csv
import zipfile
import fitz  # PyMuPDF
import streamlit as st
import pandas as pd
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Set page configuration with a premium look
st.set_page_config(
    page_title="KeyPDF Highlight & Analytics Studio",
    page_icon="🖊️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Curated High-Contrast 20 Color Palette for Multi-color Highlighting
COLOR_PALETTE = [
    ((1.0, 1.0, 0.0), "#FFFF00", "Neon Yellow"),
    ((0.2, 0.9, 0.4), "#33E666", "Mint Green"),
    ((0.0, 0.8, 1.0), "#00CCFF", "Bright Cyan"),
    ((1.0, 0.4, 0.7), "#FF66B2", "Hot Pink"),
    ((1.0, 0.6, 0.0), "#FF9900", "Mandarin Orange"),
    ((0.75, 0.52, 0.99), "#C084FC", "Lavender"),
    ((1.0, 0.42, 0.42), "#FF6B6B", "Coral Red"),
    ((0.98, 0.8, 0.08), "#FACC15", "Rich Gold"),
    ((0.65, 0.93, 0.15), "#A3E635", "Lime Green"),
    ((0.18, 0.83, 0.78), "#2DD4BF", "Turquoise"),
    ((0.53, 0.43, 0.96), "#818CF8", "Violet Sky"),
    ((0.98, 0.45, 0.63), "#FB7185", "Rose Pink"),
    ((0.99, 0.58, 0.43), "#FB923C", "Peach Coral"),
    ((0.06, 0.73, 0.51), "#10B981", "Emerald"),
    ((0.23, 0.66, 0.96), "#38BDF8", "Sky Blue"),
    ((0.91, 0.3, 0.88), "#E879F9", "Magenta"),
    ((0.96, 0.72, 0.07), "#F59E0B", "Amber"),
    ((0.96, 0.26, 0.45), "#F43F5E", "Raspberry"),
    ((0.3, 0.85, 0.39), "#4ADE80", "Spring Green"),
    ((0.02, 0.71, 0.83), "#06B6D4", "Deep Cyan")
]

# Premium Theme and Custom Styling
st.markdown("""
    <style>
    /* Main app layout */
    .stApp {
        background-color: #0F172A;
        color: #E2E8F0;
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
    }
    
    /* Sidebar styling */
    section[data-testid="stSidebar"] {
        background-color: #1E293B !important;
        border-right: 1px solid #334155;
    }
    
    /* Headers styling */
    h1, h2, h3 {
        color: #F8FAFC !important;
        font-weight: 700 !important;
    }
    
    /* Title gradient */
    .title-gradient {
        background: linear-gradient(135deg, #38BDF8 0%, #34D399 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        font-size: 2.5rem;
        font-weight: 800;
        margin-bottom: 0.5rem;
        text-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    
    /* Card containers */
    .glass-card {
        background: rgba(30, 41, 59, 0.7);
        backdrop-filter: blur(12px);
        border: 1px solid rgba(255, 255, 255, 0.08);
        border-radius: 12px;
        padding: 24px;
        margin-bottom: 20px;
        box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.2);
    }
    
    /* Highlight summary styles */
    .metric-card {
        background: rgba(15, 23, 42, 0.6);
        border-left: 4px solid #34D399;
        border-radius: 6px;
        padding: 12px 20px;
        margin: 8px 0;
    }
    
    /* Highlight labels */
    .highlight-badge {
        display: inline-block;
        padding: 2px 8px;
        border-radius: 4px;
        font-weight: 600;
        font-size: 0.85em;
        background-color: rgba(52, 211, 153, 0.15);
        color: #34D399;
        border: 1px solid rgba(52, 211, 153, 0.3);
    }
    
    /* Styled buttons */
    div.stButton > button {
        background: linear-gradient(135deg, #059669 0%, #10B981 100%) !important;
        color: white !important;
        border: none !important;
        padding: 10px 24px !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
        transition: all 0.3s ease !important;
        box-shadow: 0 4px 12px rgba(5, 150, 105, 0.3) !important;
    }
    div.stButton > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 16px rgba(5, 150, 105, 0.4) !important;
    }
    
    /* Styled file uploader borders */
    div[data-testid="stFileUploader"] {
        border: 2px dashed #475569 !important;
        padding: 16px !important;
        border-radius: 10px !important;
        background-color: #1E293B !important;
    }
    
    /* Footer styling */
    .footer {
        text-align: center;
        margin-top: 50px;
        font-size: 0.85rem;
        color: #64748B;
        border-top: 1px solid #334155;
        padding-top: 20px;
    }
    </style>
""", unsafe_allow_html=True)


def load_keywords_from_bytes(file_bytes, filename):
    """
    Parses keywords from uploaded file bytes (TXT or CSV) using various encodings.
    """
    keywords = []
    content = ""
    
    encodings = ['utf-8-sig', 'utf-8', 'cp1252', 'tis-620', 'iso-8859-11']
    for enc in encodings:
        try:
            content = file_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        content = file_bytes.decode('utf-8', errors='ignore')

    if filename.lower().endswith('.csv'):
        reader = csv.reader(io.StringIO(content))
        for row in reader:
            for cell in row:
                word = cell.strip()
                if word and word not in keywords:
                    keywords.append(word)
    else:
        for line in content.splitlines():
            word = line.strip()
            if word and word not in keywords:
                keywords.append(word)
                
    return keywords


def get_hex_color(color_name):
    """
    Returns float RGB values (0.0 to 1.0) and hex code for PyMuPDF highlighting.
    """
    for rgb, hex_val, name in COLOR_PALETTE:
        if color_name == name:
            return rgb, hex_val
    return COLOR_PALETTE[0][0], COLOR_PALETTE[0][1]


def get_pdf_fonts():
    """
    Finds standard Unicode fonts for PyMuPDF that support both Thai and English properly.
    Prioritizes bundled fonts inside the repository (fonts/ folder), then falls back to system paths.
    """
    base_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        (os.path.join(base_dir, "fonts", "leelawad.ttf"), os.path.join(base_dir, "fonts", "leelawdb.ttf")),
        (os.path.join(base_dir, "fonts", "tahoma.ttf"), os.path.join(base_dir, "fonts", "tahomabd.ttf")),
        ("C:/Windows/Fonts/leelawad.ttf", "C:/Windows/Fonts/leelawdb.ttf"),
        ("C:/Windows/Fonts/LeelawUI.ttf", "C:/Windows/Fonts/LeelaUIb.ttf"),
        ("C:/Windows/Fonts/tahoma.ttf", "C:/Windows/Fonts/tahomabd.ttf"),
        ("C:/Windows/Fonts/arial.ttf", "C:/Windows/Fonts/arialbd.ttf"),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf")
    ]
    for reg, bd in candidates:
        if os.path.exists(reg):
            bd_path = bd if os.path.exists(bd) else reg
            return reg, bd_path
    return None, None


def append_summary_pages(doc, pdf_name, summary_data, report_data, total_matches, total_pdf_words, kw_color_map):
    """
    Appends cleanly styled summary page(s) at the end of the PDF with perfect Thai & English rendering.
    Uses registered font resources without inline subset corruption.
    """
    font_reg, font_bd = get_pdf_fonts()
    original_pages = len(doc)
    
    # Create the primary summary page (Standard A4: 595 x 842 pt)
    page = doc.new_page(width=595, height=842)
    
    # Register font once per page
    if font_reg:
        page.insert_font(fontname="th_reg", fontfile=font_reg)
        page.insert_font(fontname="th_bd", fontfile=font_bd)
        fn_reg = "th_reg"
        fn_bd = "th_bd"
    else:
        fn_reg = "helv"
        fn_bd = "helv"
    
    # 1. Header Banner (Dark Slate with emerald accent line)
    page.draw_rect(fitz.Rect(30, 25, 565, 85), color=None, fill=(0.06, 0.09, 0.16))
    page.draw_rect(fitz.Rect(30, 83, 565, 85), color=None, fill=(0.2, 0.8, 0.6))
    
    page.insert_text((45, 52), "สรุปผลการวิเคราะห์และไฮไลต์คำสำคัญ (Keyword Analytics Summary)", 
                     fontsize=12, fontname=fn_bd, color=(1, 1, 1))
    page.insert_text((45, 72), "KeyPDF Highlight & Analytics Studio | เอกสารสรุปผลการตรวจจับอัตโนมัติ", 
                     fontsize=8.5, fontname=fn_reg, color=(0.8, 0.85, 0.9))
    
    # 2. Overview Metadata Card
    page.draw_rect(fitz.Rect(30, 95, 565, 155), color=(0.82, 0.86, 0.9), fill=(0.96, 0.97, 0.99))
    
    # Line 1: File name & Original Pages
    page.insert_text((45, 116), "ชื่อเอกสาร:", fontsize=9.5, fontname=fn_bd, color=(0.2, 0.25, 0.3))
    page.insert_text((115, 116), str(pdf_name)[:50], fontsize=9.5, fontname=fn_reg, color=(0.1, 0.1, 0.1))
    page.insert_text((420, 116), f"จำนวนหน้าเดิม: {original_pages} หน้า", fontsize=9, fontname=fn_reg, color=(0.3, 0.35, 0.4))
    
    # Line 2: Metrics
    overall_density = (total_matches / total_pdf_words * 100) if total_pdf_words > 0 else 0.0
    page.insert_text((45, 140), f"คำทั้งหมดใน PDF: {total_pdf_words:,} คำ", fontsize=9, fontname=fn_reg, color=(0.25, 0.3, 0.35))
    page.insert_text((220, 140), f"คำสำคัญที่พบ: {total_matches:,} จุด", fontsize=9, fontname=fn_bd, color=(0.05, 0.55, 0.35))
    page.insert_text((380, 140), f"สัดส่วนคำสำคัญ: {overall_density:.4f}%", fontsize=9, fontname=fn_reg, color=(0.25, 0.3, 0.35))
    
    # 3. Section Title
    page.insert_text((30, 173), "ตารางแจกแจงสถิติคำสำคัญและสีไฮไลต์ (Keyword Breakdown)", 
                     fontsize=10.5, fontname=fn_bd, color=(0.1, 0.15, 0.25))
    
    # 4. Table Header
    y_table = 183
    page.draw_rect(fitz.Rect(30, y_table, 565, y_table + 22), color=None, fill=(0.12, 0.16, 0.23))
    
    cols = [
        ("ลำดับ", 38),
        ("คีย์เวิร์ด (Keyword)", 65),
        ("สีไฮไลต์", 195),
        ("ชื่อเฉดสี (Color)", 240),
        ("จำนวนที่พบ", 370),
        ("% ใน PDF", 435),
        ("% คำสำคัญ", 495)
    ]
    for title, x in cols:
        page.insert_text((x, y_table + 15), title, fontsize=8, fontname=fn_bd, color=(1, 1, 1))
        
    cur_y = y_table + 22
    row_h = 22
    
    for idx, item in enumerate(summary_data):
        # Handle overflow pagination if many keywords
        if idx > 0 and idx % 25 == 0:
            page.insert_text((45, 820), f"หน้าสรุปผลการวิเคราะห์ท้ายเอกสาร | {pdf_name}", fontsize=8, fontname=fn_reg, color=(0.5, 0.5, 0.6))
            page = doc.new_page(width=595, height=842)
            if font_reg:
                page.insert_font(fontname="th_reg", fontfile=font_reg)
                page.insert_font(fontname="th_bd", fontfile=font_bd)
            cur_y = 40
            page.draw_rect(fitz.Rect(30, cur_y, 565, cur_y + 22), color=None, fill=(0.12, 0.16, 0.23))
            for title, x in cols:
                page.insert_text((x, cur_y + 15), title, fontsize=8, fontname=fn_bd, color=(1, 1, 1))
            cur_y += 22
            
        kw = item['คีย์เวิร์ด (Keyword)']
        rgb = kw_color_map.get(kw, {}).get('rgb', (1.0, 1.0, 0.0))
        color_name = item['ชื่อเฉดสี (Color Name)']
        cnt = item['จำนวนคำที่พบ (Matches)']
        p_pdf = item['% เทียบคำทั้งหมดใน PDF (% of PDF Words)']
        p_kw = item['% เทียบคำสำคัญที่พบทั้งหมด (% of Matches)']
        
        bg_col = (0.97, 0.98, 1.0) if idx % 2 == 1 else (1.0, 1.0, 1.0)
        page.draw_rect(fitz.Rect(30, cur_y, 565, cur_y + row_h), color=(0.9, 0.92, 0.95), fill=bg_col, width=0.5)
        
        page.insert_text((42, cur_y + 15), str(idx + 1), fontsize=8, fontname=fn_reg, color=(0.4, 0.4, 0.4))
        page.insert_text((65, cur_y + 15), str(kw)[:25], fontsize=8.5, fontname=fn_bd, color=(0.1, 0.1, 0.1))
        
        # Color Swatch Box
        page.draw_rect(fitz.Rect(195, cur_y + 5, 230, cur_y + 17), color=(0.6, 0.65, 0.7), fill=rgb, width=0.5)
        
        page.insert_text((240, cur_y + 15), str(color_name)[:35], fontsize=8, fontname=fn_reg, color=(0.25, 0.25, 0.25))
        page.insert_text((380, cur_y + 15), str(cnt), fontsize=8.5, fontname=fn_bd, color=(0.05, 0.45, 0.25))
        page.insert_text((435, cur_y + 15), str(p_pdf), fontsize=8, fontname=fn_reg, color=(0.3, 0.3, 0.3))
        page.insert_text((495, cur_y + 15), str(p_kw), fontsize=8, fontname=fn_reg, color=(0.3, 0.3, 0.3))
        
        cur_y += row_h

    # Footer
    page.insert_text((45, 820), f"หน้าสรุปผลการวิเคราะห์ท้ายเอกสาร | KeyPDF Highlight & Analytics Studio | {pdf_name}", 
                     fontsize=8, fontname=fn_reg, color=(0.5, 0.5, 0.6))


def process_pdf(pdf_bytes, keywords, search_flags, single_color_rgb, case_sensitive, match_whole_word=True, color_mode="อัตโนมัติ (แยกสีตามคีย์เวิร์ด)", pdf_name="document.pdf", append_summary=True, progress_callback=None):
    """
    Performs native search and highlighting using PyMuPDF.
    Supports exact Keyword matching, 20 Multi-color palette per keyword, Co-occurrence analysis,
    appends a dedicated summary page at the end of the PDF, and generates previews.
    """
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    
    # Color assignment mapping
    kw_color_map = {}
    for idx, kw in enumerate(keywords):
        if color_mode == "อัตโนมัติ (แยกสีตามคีย์เวิร์ด)":
            rgb, hex_val, name = COLOR_PALETTE[idx % len(COLOR_PALETTE)]
            kw_color_map[kw] = {"rgb": rgb, "hex": hex_val, "name": name}
        else:
            kw_color_map[kw] = {"rgb": single_color_rgb, "hex": "#FFFF00", "name": "Single Color"}
            
    summary_data = []
    report_data = []
    total_matches = 0
    total_pdf_words = 0
    keyword_match_counts = {kw: 0 for kw in keywords}
    
    total_pages = len(doc)
    page_kw_map = {p: set() for p in range(1, total_pages + 1)}
    
    for page_num in range(total_pages):
        if progress_callback:
            progress_callback(page_num + 1, total_pages)
            
        page = doc[page_num]
        page_1based = page_num + 1
        
        textpage = page.get_textpage(flags=search_flags)
        words = page.get_text("words")
        total_pdf_words += len(words)
        
        # 1. Map words to lines and build line text context
        line_coords = {}
        line_words = {}
        for w in words:
            key = (w[5], w[6])  # (block_no, line_no)
            y0 = w[1]
            if key not in line_coords:
                line_coords[key] = []
                line_words[key] = []
            line_coords[key].append(y0)
            line_words[key].append(w)
        
        line_avg_y = {key: sum(y0s)/len(y0s) for key, y0s in line_coords.items()}
        sorted_lines = sorted(line_avg_y.items(), key=lambda item: item[1])
        line_map = {key: idx + 1 for idx, (key, _) in enumerate(sorted_lines)}
        
        line_text_map = {}
        for key, ws in line_words.items():
            ws_sorted = sorted(ws, key=lambda x: x[7])
            line_text_map[key] = " ".join(x[4] for x in ws_sorted)
            
        # 2. Native Keyword Search
        for kw in keywords:
            kw_clean = kw.strip()
            if not kw_clean:
                continue
                
            rects = page.search_for(kw_clean, textpage=textpage)
            
            if rects:
                if case_sensitive:
                    rects = [r for r in rects if page.get_text("text", clip=r).strip() == kw_clean]
                
                # Filter whole word matches if requested (prevents AI inside Airport / thailand)
                if match_whole_word and rects:
                    valid_rects = []
                    is_english_kw = bool(re.match(r'^[a-zA-Z0-9_\s]+$', kw_clean))
                    
                    for rect in rects:
                        overlapping = [w for w in words if not (rect & fitz.Rect(w[:4])).is_empty]
                        if overlapping:
                            word_strs = [w[4] for w in overlapping]
                            full_word_context = " ".join(word_strs)
                            
                            if is_english_kw:
                                pattern = r'(?<![a-zA-Z0-9_])' + re.escape(kw_clean) + r'(?![a-zA-Z0-9_])'
                            else:
                                pattern = re.escape(kw_clean)
                                
                            flags = 0 if case_sensitive else re.IGNORECASE
                            if re.search(pattern, full_word_context, flags):
                                valid_rects.append(rect)
                        else:
                            valid_rects.append(rect)
                    rects = valid_rects
                
                if rects:
                    count_on_page = len(rects)
                    total_matches += count_on_page
                    keyword_match_counts[kw] += count_on_page
                    page_kw_map[page_1based].add(kw)
                    
                    highlight_rgb = kw_color_map[kw]["rgb"]
                    
                    for rect in rects:
                        annot = page.add_highlight_annot(rect)
                        if annot:
                            annot.set_colors(stroke=highlight_rgb)
                            annot.update()
                            
                        matched_block_line = None
                        for w in words:
                            w_rect = fitz.Rect(w[:4])
                            intersect = rect & w_rect
                            if not intersect.is_empty and intersect.get_area() / w_rect.get_area() > 0.5:
                                matched_block_line = (w[5], w[6])
                                break
                        
                        global_line_no = "-"
                        context_text = "-"
                        if matched_block_line:
                            global_line_no = line_map.get(matched_block_line, "-")
                            context_text = line_text_map.get(matched_block_line, "-")
                            
                        report_data.append({
                            "Keyword": kw,
                            "Page": page_1based,
                            "Line": global_line_no,
                            "Context": context_text
                        })
                        
    # Calculate Co-occurrence Matrix (Shared Pages)
    co_occurrence = {}
    for kw1 in keywords:
        co_occurrence[kw1] = {}
        for kw2 in keywords:
            shared_pages = sum(1 for p, kws in page_kw_map.items() if kw1 in kws and kw2 in kws)
            co_occurrence[kw1][kw2] = shared_pages

    # Build Summary List
    for kw, count in keyword_match_counts.items():
        pct_of_pdf = (count / total_pdf_words * 100) if total_pdf_words > 0 else 0.0
        pct_of_matches = (count / total_matches * 100) if total_matches > 0 else 0.0
        
        summary_data.append({
            "คีย์เวิร์ด (Keyword)": kw,
            "สีไฮไลต์ (Color)": kw_color_map[kw]["hex"],
            "ชื่อเฉดสี (Color Name)": kw_color_map[kw]["name"],
            "จำนวนคำที่พบ (Matches)": count,
            "% เทียบคำทั้งหมดใน PDF (% of PDF Words)": f"{pct_of_pdf:.4f}%",
            "% เทียบคำสำคัญที่พบทั้งหมด (% of Matches)": f"{pct_of_matches:.2f}%"
        })
        
    # Append Summary Page(s) at the end of the PDF
    if append_summary:
        append_summary_pages(doc, pdf_name, summary_data, report_data, total_matches, total_pdf_words, kw_color_map)

    # Save highlighted PDF
    output_buffer = io.BytesIO()
    doc.save(output_buffer, garbage=4, deflate=True)
    highlighted_pdf_bytes = output_buffer.getvalue()
    
    # Generate Preview Images
    preview_images = []
    preview_limit = min(8, len(doc))
    for page_num in range(preview_limit):
        page = doc[page_num]
        pix = page.get_pixmap(dpi=150)
        img_data = pix.tobytes("png")
        preview_images.append((page_num + 1, Image.open(io.BytesIO(img_data))))
        
    doc.close()
    
    return highlighted_pdf_bytes, summary_data, report_data, co_occurrence, total_matches, total_pdf_words, preview_images, kw_color_map


def sanitize_sheet_title(title, existing_titles):
    """
    Sanitizes sheet title for Excel (max 31 chars, no forbidden symbols, unique).
    """
    base = re.sub(r'[\\/*?:\[\]]', '_', title).strip()
    if not base:
        base = 'Sheet'
    base = base[:26]
    candidate = base
    counter = 1
    while candidate.lower() in [t.lower() for t in existing_titles]:
        candidate = f"{base[:23]}_{counter}"
        counter += 1
    existing_titles.append(candidate)
    return candidate


def generate_multi_pdf_excel(all_results):
    """
    Generates a rich Excel workbook (.xlsx) with:
    1. 'Overview (ภาพรวมทุกไฟล์)' sheet summarizing all PDF documents.
    2. Dedicated individual sheet for each PDF named cleanly after the PDF.
    """
    wb = openpyxl.Workbook()
    existing_sheets = []
    
    title_font = Font(name="Calibri", size=14, bold=True, color="0F172A")
    section_font = Font(name="Calibri", size=12, bold=True, color="1E293B")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    # 1. Overview Sheet
    ws_ov = wb.active
    ws_ov.title = "Overview (ภาพรวมทุกไฟล์)"
    existing_sheets.append(ws_ov.title)
    
    ws_ov.append(["PDF KEYWORD ANALYTICS - BATCH OVERVIEW (สรุปภาพรวมทุกไฟล์)"])
    ws_ov["A1"].font = title_font
    ws_ov.append([])
    
    ws_ov.append(["No.", "Document Name (ชื่อไฟล์)", "Total Words (คำใน PDF)", "Total Matches (คำสำคัญที่พบ)", "Keyword Density (%)", "Status"])
    for cell in ws_ov[3]:
        cell.fill = header_fill
        cell.font = header_font
        
    grand_words = 0
    grand_matches = 0
    
    for idx, res in enumerate(all_results):
        fn = res["pdf_name"]
        words = res["total_pdf_words"]
        matches = res["total_matches"]
        grand_words += words
        grand_matches += matches
        density = (matches / words * 100) if words > 0 else 0.0
        
        ws_ov.append([
            idx + 1,
            fn,
            words,
            matches,
            f"{density:.4f}%",
            "Completed"
        ])
        
    total_row = ws_ov.max_row + 1
    grand_density = (grand_matches / grand_words * 100) if grand_words > 0 else 0.0
    ws_ov.append(["", "รวมทั้งหมด (Grand Total)", grand_words, grand_matches, f"{grand_density:.4f}%", ""])
    for cell in ws_ov[total_row]:
        cell.font = Font(name="Calibri", size=11, bold=True, color="0F172A")
        cell.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        
    # 2. Individual PDF Sheets
    for res in all_results:
        pname = res["pdf_name"]
        sheet_title = sanitize_sheet_title(os.path.splitext(pname)[0], existing_sheets)
        ws = wb.create_sheet(title=sheet_title)
        
        ws.append([f"ANALYTICS SUMMARY: {pname}"])
        ws["A1"].font = title_font
        ws.append([])
        
        words = res["total_pdf_words"]
        matches = res["total_matches"]
        density = (matches / words * 100) if words > 0 else 0.0
        
        ws.append(["Metric Overview", "Value"])
        for cell in ws[3]:
            cell.fill = header_fill
            cell.font = header_font
            
        ws.append(["Document Name", pname])
        ws.append(["Total PDF Words", words])
        ws.append(["Total Keyword Matches Found", matches])
        ws.append(["Overall Keyword Density", f"{density:.4f}%"])
        ws.append([])
        
        # Keyword Breakdown Table
        ws.append(["--- KEYWORD BREAKDOWN ---"])
        ws[f"A{ws.max_row}"].font = section_font
        
        ws.append(["Keyword", "Color Name", "Color Hex", "Matches Found", "% of Total PDF Words", "% of Total Keyword Matches"])
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            
        for item in res["summary"]:
            ws.append([
                item["คีย์เวิร์ด (Keyword)"],
                item["ชื่อเฉดสี (Color Name)"],
                item["สีไฮไลต์ (Color)"],
                item["จำนวนคำที่พบ (Matches)"],
                item["% เทียบคำทั้งหมดใน PDF (% of PDF Words)"],
                item["% เทียบคำสำคัญที่พบทั้งหมด (% of Matches)"]
            ])
            
        ws.append([])
        
        # Detailed Findings Table
        ws.append(["--- DETAILED FINDINGS & POSITIONS ---"])
        ws[f"A{ws.max_row}"].font = section_font
        
        ws.append(["Keyword", "Page Number", "Line Number", "Context Line Text"])
        findings_header_row = ws.max_row
        for cell in ws[findings_header_row]:
            cell.fill = header_fill
            cell.font = header_font
            
        for row in res["report_data"]:
            ws.append([row["Keyword"], row["Page"], row["Line"], row["Context"]])
            
        ws.append([])
        
        # Co-occurrence Matrix
        if res.get("co_occurrence"):
            ws.append(["--- KEYWORD CO-OCCURRENCE MATRIX (SHARED PAGES) ---"])
            ws[f"A{ws.max_row}"].font = section_font
            
            co_occ = res["co_occurrence"]
            kws = list(co_occ.keys())
            ws.append(["Keyword"] + kws)
            co_header_row = ws.max_row
            for cell in ws[co_header_row]:
                cell.fill = header_fill
                cell.font = header_font
                
            for kw1 in kws:
                row_vals = [kw1] + [co_occ[kw1][kw2] for kw2 in kws]
                ws.append(row_vals)
                
    # Auto-adjust column widths
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 80)
            
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def generate_excel_bytes(summary, report_data, co_occurrence, total_pdf_words, total_matches, pdf_name="Document"):
    """
    Convenience wrapper for single-document Excel generation.
    """
    single_res = {
        "pdf_name": pdf_name,
        "summary": summary,
        "report_data": report_data,
        "co_occurrence": co_occurrence,
        "total_pdf_words": total_pdf_words,
        "total_matches": total_matches
    }
    return generate_multi_pdf_excel([single_res])


def generate_csv_bytes(summary, report_data, total_pdf_words, total_matches):
    """
    Generates standard CSV content with UTF-8 BOM encoding.
    """
    output = io.StringIO()
    writer = csv.writer(output)
    
    overall_density = (total_matches / total_pdf_words * 100) if total_pdf_words > 0 else 0.0
    
    writer.writerow(["--- PDF KEYWORD SEARCH SUMMARY ---"])
    writer.writerow(["Total PDF Words", total_pdf_words])
    writer.writerow(["Total Keyword Matches", total_matches])
    writer.writerow(["Overall Density", f"{overall_density:.4f}%"])
    writer.writerow([])
    
    writer.writerow(["--- KEYWORD BREAKDOWN ---"])
    writer.writerow(["Keyword", "Color Name", "Color Hex", "Matches Count", "% of Total PDF Words", "% of Total Keyword Matches"])
    for item in summary:
        writer.writerow([
            item["คีย์เวิร์ด (Keyword)"],
            item["ชื่อเฉดสี (Color Name)"],
            item["สีไฮไลต์ (Color)"],
            item["จำนวนคำที่พบ (Matches)"],
            item["% เทียบคำทั้งหมดใน PDF (% of PDF Words)"],
            item["% เทียบคำสำคัญที่พบทั้งหมด (% of Matches)"]
        ])
        
    writer.writerow([])
    
    writer.writerow(["--- DETAILED FINDINGS ---"])
    writer.writerow(["Keyword", "Page Number", "Line Number", "Context Line Text"])
    for row in report_data:
        writer.writerow([row["Keyword"], row["Page"], row["Line"], row["Context"]])
        
    return output.getvalue().encode('utf-8-sig')


def generate_batch_zip(all_results):
    """
    Packages all highlighted PDFs (with attached summary pages) into a single ZIP archive.
    """
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for res in all_results:
            pname = os.path.splitext(res["pdf_name"])[0]
            # Write PDF (with appended summary page)
            zf.writestr(f"{pname}_highlighted.pdf", res["pdf_bytes"])
    return zip_buf.getvalue()


# --- STREAMLIT UI ---

def main():
    st.markdown('<div class="title-gradient">KeyPDF Highlight & Analytics Studio</div>', unsafe_allow_html=True)
    st.markdown("<p style='color: #94A3B8; font-size: 1.1rem; margin-top:-10px;'>ยกระดับการตรวจทานเอกสาร PDF ค้นหาตรงตามคำเต็ม แยกสีไฮไลต์อัจฉริยะ 20 เฉดสี พร้อมแนบหน้าสรุปท้ายเล่มและส่งออก Excel แยก Sheet</p>", unsafe_allow_html=True)
    st.write("---")
    
    st.sidebar.markdown("### 🛠️ อัปโหลดและตั้งค่า (Upload & Settings)")
    
    # 1. Upload PDF (Supports Multiple Files)
    uploaded_pdfs = st.sidebar.file_uploader(
        "1. เลือกไฟล์ PDF ที่ต้องการไฮไลต์ (รองรับหลายไฟล์พร้อมกัน)", 
        type=["pdf"],
        accept_multiple_files=True,
        help="สามารถเลือกหรือลากวางไฟล์ PDF ได้ทีละหลายไฟล์พร้อมกัน (รองรับขนาดไฟล์สูงสุด 1,024 MB หรือ 1 GB ต่อไฟล์)"
    )
    
    if uploaded_pdfs:
        st.sidebar.markdown(f"<span class='highlight-badge'>อัปโหลด PDF: {len(uploaded_pdfs)} ไฟล์</span>", unsafe_allow_html=True)
    
    # 2. Upload/Input Keywords
    st.sidebar.write("")
    st.sidebar.markdown("**2. ระบุคีย์เวิร์ดคำสำคัญ**")
    kw_source = st.sidebar.radio(
        "วิธีการใส่คีย์เวิร์ด:",
        ["อัปโหลดไฟล์ (.txt, .csv)", "พิมพ์คีย์เวิร์ดด้วยตนเอง"],
        label_visibility="collapsed"
    )
    
    keywords = []
    if kw_source == "อัปโหลดไฟล์ (.txt, .csv)":
        uploaded_kw = st.sidebar.file_uploader(
            "อัปโหลดไฟล์คีย์เวิร์ด (.txt / .csv)", 
            type=["txt", "csv"],
            help="ไฟล์ข้อความธรรมดา (หนึ่งคำต่อบรรทัด) หรือไฟล์ CSV"
        )
        if uploaded_kw:
            keywords = load_keywords_from_bytes(uploaded_kw.read(), uploaded_kw.name)
    else:
        text_kw = st.sidebar.text_area(
            "พิมพ์คำสำคัญที่ต้องการ (หนึ่งคำต่อหนึ่งบรรทัด):",
            placeholder="ตัวอย่างเช่น:\nAI\nPython\nสัญญาจ้าง\nบริษัท",
            height=130
        )
        if text_kw:
            keywords = [line.strip() for line in text_kw.splitlines() if line.strip()]

    if keywords:
        st.sidebar.markdown(f"<span class='highlight-badge'>โหลดคีย์เวิร์ดสำเร็จ: {len(keywords)} คำ</span>", unsafe_allow_html=True)
    
    st.sidebar.write("")
    st.sidebar.markdown("### 🎨 ตั้งค่าการค้นหาและสีไฮไลต์ (20 เฉดสี)")
    
    # Match Whole Word option
    match_whole_word = st.sidebar.checkbox(
        "ตรงตามคำเต็มเท่านั้น (Match Whole Word Only)",
        value=True,
        help="หากเปิดใช้งาน คำว่า 'AI' จะไม่นับคำที่มี ai อยู่ภายใน เช่น 'Airport' หรือ 'thailand'"
    )
    
    # Color Modes
    color_mode = st.sidebar.radio(
        "โหมดสีของไฮไลต์ (Highlight Color Mode):",
        ["อัตโนมัติ (แยกสีตามคีย์เวิร์ด)", "สีเดี่ยว (Single Color)"],
        index=0
    )
    
    single_color_rgb = (1.0, 1.0, 0.0)
    single_color_choice = COLOR_PALETTE[0][2]
    if color_mode == "สีเดี่ยว (Single Color)":
        single_color_choice = st.sidebar.selectbox(
            "เลือกสีไฮไลต์เดี่ยว (20 เฉดสี):",
            [name for _, _, name in COLOR_PALETTE],
            index=0
        )
        single_color_rgb, _ = get_hex_color(single_color_choice)
    
    case_sensitive = st.sidebar.checkbox(
        "Case Sensitive (ตัวพิมพ์เล็ก-ใหญ่)", 
        value=False
    )
    
    dehyphenate = st.sidebar.checkbox(
        "Dehyphenate (รวมคำตัดยติภังค์ท้ายบรรทัด)", 
        value=True
    )
    
    append_summary_page = st.sidebar.checkbox(
        "📄 แนบหน้าสรุปผลต่อท้าย PDF อัตโนมัติ",
        value=True,
        help="เพิ่มหน้าสรุปสถิติและตารางสีไฮไลต์เข้าไปที่หน้าสุดท้ายของแต่ละไฟล์ PDF ทันที"
    )
    
    search_flags = 2 | 1
    if dehyphenate:
        search_flags |= 16
    
    # Main panel
    col1, col2 = st.columns([1.1, 1.1])
    
    with col1:
        st.markdown('<div class="glass-card">', unsafe_allow_html=True)
        st.subheader("📋 รายละเอียดการประมวลผล")
        
        if not uploaded_pdfs:
            st.info("👋 กรุณาอัปโหลดไฟล์ PDF (หนึ่งไฟล์หรือหลายไฟล์) ในแผงควบคุมด้านซ้ายเพื่อเริ่มต้น")
        elif not keywords:
            st.warning("⚠️ กรุณาระบุหรืออัปโหลดคีย์เวิร์ดคำสำคัญ")
        else:
            file_names = [f.name for f in uploaded_pdfs]
            st.write(f"📂 **ไฟล์ PDF ที่เลือก:** `{len(uploaded_pdfs)} ไฟล์` ({', '.join(file_names[:3])}{' ...' if len(file_names) > 3 else ''})")
            st.write(f"🔑 **จำนวนคีย์เวิร์ด:** `{len(keywords)} คำ`")
            st.write(f"🎯 **โหมดการค้นหา:** `{'ตรงตามคำเต็มเท่านั้น (Match Whole Word)' if match_whole_word else 'รวมคำส่วนย่อย (Substring Match)'}`")
            st.write(f"🎨 **โหมดสี:** `{color_mode}`")
            st.write(f"📄 **แนบหน้าสรุปท้าย PDF:** `{'เปิดใช้งาน (แนบหน้าสรุปท้ายเล่ม)' if append_summary_page else 'ปิดใช้งาน'}`")
            
            run_btn = st.button("🚀 เริ่มไฮไลต์และวิเคราะห์เอกสารทั้งหมด (Batch Process)")
            
            if run_btn:
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                all_results = []
                total_files = len(uploaded_pdfs)
                
                try:
                    for f_idx, up_file in enumerate(uploaded_pdfs):
                        pdf_name = up_file.name
                        pdf_bytes = up_file.read()
                        
                        def update_progress(curr_page, total_pages, current_file_idx=f_idx, current_filename=pdf_name):
                            overall = (current_file_idx + (curr_page / total_pages)) / total_files
                            progress_bar.progress(min(max(overall, 0.0), 1.0))
                            status_text.text(f"⏳ ไฟล์ {current_file_idx + 1}/{total_files}: {current_filename} (หน้า {curr_page}/{total_pages})...")
                            
                        res_bytes, summary, report_data, co_occurrence, count, total_pdf_words, previews, kw_color_map = process_pdf(
                            pdf_bytes, 
                            keywords, 
                            search_flags, 
                            single_color_rgb, 
                            case_sensitive,
                            match_whole_word,
                            color_mode,
                            pdf_name=pdf_name,
                            append_summary=append_summary_page,
                            progress_callback=update_progress
                        )
                        
                        all_results.append({
                            "pdf_name": pdf_name,
                            "pdf_bytes": res_bytes,
                            "summary": summary,
                            "report_data": report_data,
                            "co_occurrence": co_occurrence,
                            "total_matches": count,
                            "total_pdf_words": total_pdf_words,
                            "previews": previews,
                            "kw_color_map": kw_color_map
                        })
                        
                    # Generate Consolidated Excel & ZIP archive
                    batch_excel_bytes = generate_multi_pdf_excel(all_results)
                    batch_zip_bytes = generate_batch_zip(all_results)
                    
                    st.session_state["all_results"] = all_results
                    st.session_state["batch_excel_bytes"] = batch_excel_bytes
                    st.session_state["batch_zip_bytes"] = batch_zip_bytes
                    st.session_state["selected_doc_idx"] = 0
                    
                    progress_bar.empty()
                    status_text.empty()
                    
                    grand_matches = sum(r["total_matches"] for r in all_results)
                    grand_words = sum(r["total_pdf_words"] for r in all_results)
                    st.success(f"✨ ประมวลผลเสร็จสิ้นครบทั้ง {total_files} ไฟล์! พบคำสำคัญรวม {grand_matches:,} จุด จากทั้งหมด {grand_words:,} คำ")
                    
                except Exception as e:
                    progress_bar.empty()
                    status_text.empty()
                    st.error(f"❌ เกิดข้อผิดพลาดในการประมวลผล PDF: {str(e)}")
                    
        # Display Results
        if "all_results" in st.session_state and st.session_state["all_results"]:
            all_results = st.session_state["all_results"]
            total_files = len(all_results)
            
            st.write("---")
            st.markdown("#### 🎉 สรุปผลลัพธ์และดาวน์โหลดเอกสาร")
            
            grand_matches = sum(r["total_matches"] for r in all_results)
            grand_words = sum(r["total_pdf_words"] for r in all_results)
            grand_density = (grand_matches / grand_words * 100) if grand_words > 0 else 0.0
            
            m_col1, m_col2, m_col3 = st.columns(3)
            with m_col1:
                st.metric("จำนวนไฟล์ทั้งหมด", f"{total_files} ไฟล์")
            with m_col2:
                st.metric("คำสำคัญที่พบรวมทุกไฟล์", f"{grand_matches:,} จุด")
            with m_col3:
                st.metric("สัดส่วนคำสำคัญรวม", f"{grand_density:.4f}%")
                
            st.write("")
            
            # Batch Downloads
            d_col1, d_col2 = st.columns(2)
            with d_col1:
                st.download_button(
                    label="📊 ดาวน์โหลด Excel สรุปทุกไฟล์ (.xlsx)",
                    data=st.session_state["batch_excel_bytes"],
                    file_name="All_PDFs_Keyword_Analytics_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="ไฟล์ Excel รวมที่มี Sheet สรุปภาพรวม และ Sheet แยกตามชื่อแต่ละ PDF",
                    use_container_width=True
                )
            with d_col2:
                st.download_button(
                    label="📦 ดาวน์โหลด PDF ทั้งหมดรวม (.ZIP)",
                    data=st.session_state["batch_zip_bytes"],
                    file_name="All_Highlighted_PDFs.zip",
                    mime="application/zip",
                    help="ไฟล์ ZIP รวมเอกสาร PDF ที่ไฮไลต์พร้อมหน้าสรุปท้ายเล่มทุกไฟล์",
                    use_container_width=True
                )
                
            st.write("---")
            
            # Document selector if multiple files
            if total_files > 1:
                doc_options = [f"📂 {idx+1}. {r['pdf_name']} (พบ {r['total_matches']:,} จุด)" for idx, r in enumerate(all_results)]
                selected_opt = st.selectbox("🔍 เลือกเอกสารที่ต้องการดูรายละเอียดและพรีวิว:", doc_options, index=0)
                selected_idx = doc_options.index(selected_opt)
            else:
                selected_idx = 0
                
            st.session_state["selected_doc_idx"] = selected_idx
            cur_res = all_results[selected_idx]
            cur_pdf_name = cur_res["pdf_name"]
            
            st.markdown(f"##### 📄 รายละเอียดเอกสาร: `{cur_pdf_name}`")
            
            doc_words = cur_res["total_pdf_words"]
            doc_matches = cur_res["total_matches"]
            doc_density = (doc_matches / doc_words * 100) if doc_words > 0 else 0.0
            
            if doc_words == 0:
                st.warning(f"⚠️ **ไม่พบข้อความ (Text Layer) ในไฟล์ {cur_pdf_name}** — เอกสารอาจเป็นภาพสแกน แนะนำให้ผ่านระบบ OCR ก่อน")
                
            sub_m1, sub_m2, sub_m3 = st.columns(3)
            with sub_m1:
                st.metric("จำนวนคำใน PDF นี้", f"{doc_words:,} คำ")
            with sub_m2:
                st.metric("คำสำคัญที่พบในไฟล์นี้", f"{doc_matches:,} จุด")
            with sub_m3:
                st.metric("สัดส่วนคำสำคัญ", f"{doc_density:.4f}%")
                
            st.write("")
            
            # Individual file download buttons (PDF & Excel only)
            ind_col1, ind_col2 = st.columns(2)
            pname_base = os.path.splitext(cur_pdf_name)[0]
            with ind_col1:
                st.download_button(
                    label="📥 ดาวน์โหลด PDF ไฮไลต์นี้",
                    data=cur_res["pdf_bytes"],
                    file_name=f"{pname_base}_highlighted.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )
            with ind_col2:
                single_excel = generate_excel_bytes(
                    cur_res["summary"], cur_res["report_data"], cur_res["co_occurrence"],
                    cur_res["total_pdf_words"], cur_res["total_matches"], pdf_name=cur_pdf_name
                )
                st.download_button(
                    label="📊 ดาวน์โหลด Excel เฉพาะไฟล์นี้ (.xlsx)",
                    data=single_excel,
                    file_name=f"{pname_base}_analytics.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
            st.write("")
            
            tab1, tab2, tab3 = st.tabs([
                "📊 ตารางแจกแจงสถิติ & สีไฮไลต์", 
                "📌 คำสำคัญที่พบร่วมกัน (Co-occurrence)", 
                "🔍 รายละเอียดตำแหน่งที่พบ (Findings)"
            ])
            
            with tab1:
                st.write(f"**ตารางวิเคราะห์คำสำคัญและสีไฮไลต์ประจำคำ ({cur_pdf_name}):**")
                df_summary = pd.DataFrame(cur_res["summary"])[
                    ["คีย์เวิร์ด (Keyword)", "ชื่อเฉดสี (Color Name)", "สีไฮไลต์ (Color)", "จำนวนคำที่พบ (Matches)", "% เทียบคำทั้งหมดใน PDF (% of PDF Words)", "% เทียบคำสำคัญที่พบทั้งหมด (% of Matches)"]
                ]
                
                def make_color_badge(hex_code):
                    return f'<span style="background-color: {hex_code}; color: #000; padding: 3px 10px; border-radius: 4px; font-weight: bold;">■ {hex_code}</span>'
                
                df_summary_display = df_summary.copy()
                df_summary_display["สีไฮไลต์ (Color)"] = df_summary_display["สีไฮไลต์ (Color)"].apply(make_color_badge)
                
                st.write(
                    df_summary_display.to_html(escape=False, index=False),
                    unsafe_allow_html=True
                )
                
            with tab2:
                st.write(f"**ตารางวิเคราะห์คำสำคัญที่พบอยู่ร่วมกันในหน้าเดียวกัน (Shared Pages Matrix):**")
                co_matrix = cur_res["co_occurrence"]
                if co_matrix:
                    df_co = pd.DataFrame(co_matrix)
                    st.dataframe(df_co, use_container_width=True)
                else:
                    st.info("ไม่พบข้อมูล Co-occurrence")
                    
            with tab3:
                st.write(f"**รายละเอียดตำแหน่งหน้า บรรทัด และข้อความบริบท ({cur_pdf_name}):**")
                df_report = pd.DataFrame(cur_res["report_data"])
                if not df_report.empty:
                    st.dataframe(df_report, use_container_width=True, hide_index=True)
                else:
                    st.info("ไม่พบคำสำคัญในเอกสารนี้")
                    
        st.markdown('</div>', unsafe_allow_html=True)
        
    with col2:
        st.markdown('<div class="glass-card">', unsafe_allow_html=True)
        st.subheader("👁️ พรีวิวหน้าเอกสาร (PDF Preview)")
        
        if "all_results" in st.session_state and st.session_state["all_results"]:
            sel_idx = st.session_state.get("selected_doc_idx", 0)
            cur_res = st.session_state["all_results"][sel_idx]
            previews = cur_res.get("previews", [])
            
            if previews:
                st.markdown(f"**กำลังแสดงตัวอย่างของ:** `{cur_res['pdf_name']}` (มีหน้าสรุปท้ายเล่ม)")
                if len(previews) > 1:
                    page_num_sel = st.slider(
                        "เลือกหน้าที่ต้องการดูตัวอย่าง:", 
                        min_value=1, 
                        max_value=len(previews), 
                        value=len(previews)  # Default to showing the summary page or page 1
                    )
                    selected_img = previews[page_num_sel - 1][1]
                    caption_text = f"ตัวอย่างหน้าที่ {page_num_sel}"
                    if page_num_sel == len(previews) and append_summary_page:
                        caption_text += " (📄 หน้าสรุปผลการวิเคราะห์ท้ายเล่ม)"
                    st.image(
                        selected_img, 
                        caption=caption_text, 
                        use_container_width=True
                    )
                else:
                    st.image(
                        previews[0][1], 
                        caption="ตัวอย่างหน้าแรก", 
                        use_container_width=True
                    )
            else:
                st.info("ไม่สามารถสร้างภาพพรีวิวสำหรับเอกสารนี้ได้")
        else:
            st.info("💡 เมื่อกดเริ่มไฮไลต์เอกสาร ระบบจะแสดงพรีวิวหน้าเอกสารพร้อมหน้าสรุปท้ายเล่มที่ตรงนี้")
            st.markdown(
                """
                <div style='text-align: center; padding: 40px; border: 2px dashed rgba(255,255,255,0.05); border-radius: 8px;'>
                    <span style='font-size: 4rem;'>📄</span>
                    <p style='color: #64748B; margin-top: 15px;'>ยังไม่มีเอกสารตัวอย่างที่จะพรีวิว</p>
                </div>
                """,
                unsafe_allow_html=True
            )
            
        st.markdown('</div>', unsafe_allow_html=True)

    # Footer
    st.markdown("""
        <div class="footer">
            KeyPDF Highlight & Analytics Studio &copy; 2026 | พัฒนาด้วย Streamlit & PyMuPDF (fitz)
        </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
